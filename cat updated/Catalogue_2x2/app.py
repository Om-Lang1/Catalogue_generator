from flask import Flask, render_template, request, send_file, jsonify
import pypyodbc as odbc
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image as RLImage, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from PIL import Image
import os
from io import BytesIO
from datetime import datetime
import glob
import pandas as pd

app = Flask(__name__)

# Database connection configuration
DRIVER_NAME = 'SQL SERVER'
SERVER_NAME = 'SJSERVER'
DATABASE_NAME = 'TSHISJEP'

def find_image_by_style_code(style_code, images_base_path="\\\\sjserver\\GatiSoftTech\\Images"):
    """Search for images in all subfolders of Images directory based on style code"""
    if not style_code or not os.path.exists(images_base_path):
        return None
    
    # Define possible image patterns to search for
    patterns = [
        f"*{style_code}*.jpg",
        f"*{style_code}*.jpeg", 
        f"*{style_code}*.png",
        f"*{style_code}*.bmp",
        f"*{style_code}*.gif"
    ]
    
    # Search recursively in all subfolders
    for pattern in patterns:
        search_path = os.path.join(images_base_path, "**", pattern)
        matching_files = glob.glob(search_path, recursive=True)
        
        if matching_files:
            # Return the first match (you could modify to return all matches or prioritize certain folders)
            return matching_files[0]
    
    return None

def get_db_connection():
    """Create database connection using Windows authentication"""
    try:
        connection_string = f"""
            DRIVER={{{DRIVER_NAME}}};
            SERVER={SERVER_NAME};
            DATABASE={DATABASE_NAME};
            Trust_Connection=yes;
        """
        conn = odbc.connect(connection_string)
        return conn
    except Exception as e:
        print(f"Database connection error: {e}")
        return None

@app.route('/')
def index():
    """Catalogue generator home page"""
    return render_template('catalogue.html')


def get_catalogue_data(style_codes):
    """Fetch catalogue data for specific style codes"""
    conn = get_db_connection()
    if conn is None:
        return None
    
    try:
        cursor = conn.cursor()
        # Convert list to comma-separated string for SQL IN clause
        style_codes_str = "','".join(style_codes)
        
        query = f"""
            WITH DiamondData AS (
                SELECT
                    a.StyleId,
                    a.StyleCode,
                    b.CategoryGroup,
                    b.Category,
                    d.RawMitName AS RmType_D,
                    d.ItemCode AS RmCode_D,
                    d.QlyCode AS QltyCode_D,
                    e.SizeCode AS SizeCode_D,
                    c.Pieces AS Pieces_D,
                    c.NetWeight AS Weight_D,
                    CASE
                        WHEN c.Pieces > 0 THEN ROUND(c.NetWeight / c.Pieces, 4)
                        ELSE 0
                    END AS AvgPointer_D,
                    f.SetCode,
                    f.SetName,
                    ROW_NUMBER() OVER (
                        PARTITION BY a.StyleId
                        ORDER BY d.RawMitName, e.SizeCode DESC, c.NetWeight
                    ) AS D_rn
                FROM StyleMst a
                INNER JOIN StyleMstCatchDetailView b
                    ON a.StyleId = b.H_StyleId AND b.H_SrNo = 1
                INNER JOIN StyleMstDetail c
                    ON a.StyleId = c.StyleId
                LEFT JOIN SPM_ItemView d
                    ON c.ItemId = d.ItemId
                LEFT JOIN SizeMst e
                    ON c.SizeNo = e.SizeNo
                LEFT JOIN SettingMst f
                    ON c.SMDSetNo = f.SetNo
                WHERE d.RawMitName LIKE '%Diamond%'
            ),

            MetalData AS (
                SELECT
                    a.StyleId,
                    a.StyleCode,
                    d.RawMitName AS RmType_M,
                    d.ItemCode AS RmCode_M,
                    d.QlyCode AS QltyCode_M,
                    e.SizeCode AS SizeCode_M,
                    c.Pieces AS Pieces_M,
                    c.NetWeight AS Weight_M,
                    CASE
                        WHEN c.Pieces > 0 THEN ROUND(c.NetWeight / c.Pieces, 4)
                        ELSE 0
                    END AS AvgPointer_M,
                    ROW_NUMBER() OVER (
                        PARTITION BY a.StyleId
                        ORDER BY d.RawMitName, d.QlyCode
                    ) AS M_rn
                FROM StyleMst a
                INNER JOIN StyleMstDetail c
                    ON a.StyleId = c.StyleId
                LEFT JOIN SPM_ItemView d
                    ON c.ItemId = d.ItemId
                LEFT JOIN SizeMst e
                    ON c.SizeNo = e.SizeNo
                WHERE d.RawMitName LIKE '%Metal%'
            ),

            ColourStoneData AS (
                SELECT
                    a.StyleId,
                    a.StyleCode,
                    d.RawMitName AS RmType_CS,
                    d.ItemCode AS RmCode_CS,
                    d.QlyCode AS QltyCode_CS,
                    e.SizeCode AS SizeCode_CS,
                    c.Pieces AS Pieces_CS,
                    c.NetWeight AS Weight_CS,
                    CASE
                        WHEN c.Pieces > 0 THEN ROUND(c.NetWeight / c.Pieces, 4)
                        ELSE 0
                    END AS AvgPointer_CS,
                    ROW_NUMBER() OVER (
                        PARTITION BY a.StyleId
                        ORDER BY d.RawMitName, d.QlyCode
                    ) AS C_rn
                FROM StyleMst a
                INNER JOIN StyleMstDetail c
                    ON a.StyleId = c.StyleId
                LEFT JOIN SPM_ItemView d
                    ON c.ItemId = d.ItemId
                LEFT JOIN SizeMst e
                    ON c.SizeNo = e.SizeNo
                WHERE d.RawMitName LIKE '%Stone%'
            ),

            CombinedData AS (
                SELECT
                    COALESCE(d.StyleId, m.StyleId, c.StyleId) AS StyleId,
                    d.CategoryGroup,
                    d.Category,
                    COALESCE(d.StyleCode, m.StyleCode, c.StyleCode) AS StyleCode,
                    d.RmType_D, d.RmCode_D, d.QltyCode_D, d.SizeCode_D,
                    d.Pieces_D, d.Weight_D, d.AvgPointer_D,
                    m.RmType_M, m.RmCode_M, m.QltyCode_M, m.SizeCode_M,
                    m.Pieces_M, m.Weight_M, m.AvgPointer_M,
                    c.RmType_CS, c.RmCode_CS, c.QltyCode_CS, c.SizeCode_CS,
                    c.Pieces_CS, c.Weight_CS, c.AvgPointer_CS,
                    d.SetCode,
                    d.SetName,
                    d.D_rn,
                    m.M_rn,
                    c.C_rn
                FROM DiamondData d
                FULL OUTER JOIN MetalData m
                    ON d.StyleId = m.StyleId AND d.D_rn = m.M_rn
                FULL OUTER JOIN ColourStoneData c
                    ON COALESCE(d.StyleId, m.StyleId) = c.StyleId
                   AND COALESCE(d.D_rn, m.M_rn) = c.C_rn
                WHERE COALESCE(d.StyleCode, m.StyleCode, c.StyleCode) IN ('{style_codes_str}')
            )

            SELECT
                cd.StyleId,
                CASE WHEN COALESCE(cd.D_rn, cd.M_rn, cd.C_rn) = 1 THEN cd.CategoryGroup END AS CategoryGroup,
                CASE WHEN COALESCE(cd.D_rn, cd.M_rn, cd.C_rn) = 1 THEN cd.Category END AS Category,
                CASE WHEN COALESCE(cd.D_rn, cd.M_rn, cd.C_rn) = 1 THEN cd.StyleCode END AS StyleCode,
                cd.Pieces_D,
                cd.Weight_D,
                cd.Weight_M,
                cd.Pieces_CS,
                cd.QltyCode_CS,
                cd.SetCode,
                cd.SetName,
                CASE WHEN COALESCE(cd.D_rn, cd.M_rn, cd.C_rn) = 1
                     THEN '\\\\sjserver\\GatiSoftTech\\Images\\' + LEFT(cd.StyleCode, 2) + '\\DM 3D ' + cd.StyleCode + '.jpg'
                     ELSE '' END AS Img3DPath,
                CASE WHEN COALESCE(cd.D_rn, cd.M_rn, cd.C_rn) = 1
                     THEN '\\\\sjserver\\GatiSoftTech\\Images\\' + LEFT(cd.StyleCode, 2) + '\\DM LD ' + cd.StyleCode + '.jpg'
                     ELSE '' END AS ImgLDPath,
                CASE WHEN COALESCE(cd.D_rn, cd.M_rn, cd.C_rn) = 1
                     THEN '\\\\sjserver\\GatiSoftTech\\Images\\' + LEFT(cd.StyleCode, 2) + '\\DM 3D ' + cd.StyleCode + '.png'
                     ELSE '' END AS ImgPngPath,
                1 AS FinalSort,
                COALESCE(cd.D_rn, cd.M_rn, cd.C_rn) AS DetailSort
            FROM CombinedData cd
            ORDER BY cd.StyleId, FinalSort, DetailSort
        """
        
        cursor.execute(query)
        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        
        # Convert column names to lowercase for consistent access
        columns_lower = [col.lower() for col in columns]
        
        # Group data by StyleCode (using StyleId to group all rows for same product)
        grouped_data = {}
        for row in rows:
            row_dict = dict(zip(columns_lower, row))
            style_id = row_dict.get('styleid')
            style_code = row_dict.get('stylecode')
            
            # Use StyleId as the grouping key since StyleCode can be None in some rows
            if style_id:
                if style_id not in grouped_data:
                    grouped_data[style_id] = {
                        'StyleCode': style_code,  # Will be set from first row (where it's not None)
                        'Pieces_D': 0,
                        'Weight_D': 0.0,
                        'Weight_M': 0.0,
                        'Pieces_CS': 0,
                        'ColorStones': {},  # Dictionary to store stone quality codes and their total pieces
                        'RmCode_CS': None,
                        'SetCode': None,
                        'SetName': None,
                        'ImagePath': row_dict.get('img3dpath') or row_dict.get('imgldpath', ''),
                        'Img3DPath': row_dict.get('img3dpath', ''),
                        'ImgLDPath': row_dict.get('imgldpath', ''),
                        'ImgPngPath': row_dict.get('imgpngpath', '')
                    }
                
                # Update StyleCode if we have one and don't have one yet
                if style_code and not grouped_data[style_id]['StyleCode']:
                    grouped_data[style_id]['StyleCode'] = style_code
                
                # Aggregate the numeric values for ALL rows of this style
                pieces_d = row_dict.get('pieces_d')
                if pieces_d is not None:
                    try:
                        grouped_data[style_id]['Pieces_D'] += int(pieces_d)
                    except (ValueError, TypeError):
                        pass
                
                weight_d = row_dict.get('weight_d')
                if weight_d is not None:
                    try:
                        grouped_data[style_id]['Weight_D'] += float(weight_d)
                    except (ValueError, TypeError):
                        pass
                
                weight_m = row_dict.get('weight_m')
                if weight_m is not None:
                    try:
                        grouped_data[style_id]['Weight_M'] += float(weight_m)
                    except (ValueError, TypeError):
                        pass
                
                # Aggregate color stones by quality code for ALL rows of this style
                qlty_code_cs = row_dict.get('qltycode_cs')
                pieces_cs = row_dict.get('pieces_cs')
                
                if qlty_code_cs and pieces_cs is not None:
                    try:
                        pieces_cs_int = int(pieces_cs)
                        if qlty_code_cs not in grouped_data[style_id]['ColorStones']:
                            grouped_data[style_id]['ColorStones'][qlty_code_cs] = 0
                        grouped_data[style_id]['ColorStones'][qlty_code_cs] += pieces_cs_int
                        grouped_data[style_id]['Pieces_CS'] += pieces_cs_int
                    except (ValueError, TypeError):
                        pass
                
                # ONLY set other fields from rows where StyleCode is present (not NULL)
                if style_code:
                    # Set the first quality code for backward compatibility
                    if qlty_code_cs and not grouped_data[style_id]['RmCode_CS']:
                        grouped_data[style_id]['RmCode_CS'] = qlty_code_cs
                        
                        # Map quality code to stone name
                        stone_name_map = {
                            'SI': 'SI',
                            'VS2SI1': 'VS2SI1',
                            'SI1SI2': 'SI1SI2',
                            'ED1-SI': 'ED1-SI',
                            'SI2': 'SI2',
                            'I3': 'I3',
                            'VS': 'VS',
                            'SI3-I1': 'SI3-I1',
                            'Treated': 'Treated',
                            'LMX': 'LMX',
                            'OWLB I2': 'OWLB I2',
                            'TTLC-I1': 'TTLC-I1',
                            'I1I2': 'I1I2',
                            'I1': 'I1',
                            'I2I3': 'I2I3',
                            'MIX': 'MIX',
                            'Broken': 'Broken',
                            'ED1-SI1-SI2': 'ED1-SI1-SI2',
                            'ED1-VS': 'ED1-VS',
                            'ED1-VS-DEF': 'ED1-VS-DEF',
                            'SHF-I2': 'SHF-I2',
                            'SHF-I1': 'SHF-I1',
                            'SP': 'Sapphire',
                            'TZ': 'Tanzanite',
                            'LONDON': 'London Topaz',
                            'OMJ-SI': 'OMJ-SI',
                            'PL': 'PEARL',
                            'SI1': 'SI1',
                            'I2NATS': 'I2NATS',
                            'EM': 'Emerald',
                            'RU': 'Ruby',
                            'ID-SI': 'ID-SI',
                            'VS2': 'VS2',
                            'IDI-SI': 'IDI-SI',
                            'IDI-VS': 'IDI-VS',
                            'HKD-VS': 'HKD-VS',
                            'SP-AA': 'Sapphire AA',
                            'ANJ-I1': 'ANJ-I1',
                            'ANJ-I3': 'ANJ-I3',
                            'GR': 'Garnet',
                            'I2WH': 'I2WH',
                            'DHI-VS': 'DHI-VS',
                            'AM': 'Amethyst',
                            'VDC-I1I2': 'VDC-I1I2',
                            'UJI-SI1': 'UJI-SI1',
                            'UJI-VS': 'UJI-VS',
                            'REJ': 'REJ',
                            'ANJ-SI': 'ANJ-SI',
                            'SGI-VS1': 'SGI-VS1',
                            'SHF-I3': 'SHF-I3',
                            'CIT': 'CITRINE',
                            'AJP-I3': 'AJP-I3',
                            'SGI-VVS2': 'SGI-VVS2',
                            'SGI-VS': 'SGI-VS',
                            'SGI-SI': 'SGI-SI',
                            'SGI-VVS': 'SGI-VVS',
                            'BDL-VS': 'BDL-VS',
                            'SGI-SI1': 'SGI-SI1',
                            'PD': 'Peridot',
                            'BDL-VS2': 'BDL-VS2',
                            'ANJ-I3NTS': 'ANJ-I3NTS',
                            'AJP-I2': 'AJP-I2',
                            'VHP-I1': 'VHP-I1',
                            'NUA-I3': 'NUA-I3',
                            'AQM': 'Aquamarine',
                            'VS-SI1': 'VS-SI1',
                            'SI3': 'SI3',
                            'DHI-SI3-I1': 'DHI-SI3-I1',
                            'SPNL': 'SPNL',
                            'I1I2-NTS': 'I1I2-NTS',
                            'NLL-SI': 'NLL-SI',
                            'NUA-I2': 'NUA-I2',
                            'IDJ-SI': 'IDJ-SI',
                            'IDJ-I2': 'IDJ-I2',
                            'SPGL': 'Sapphire Glass Filled',
                            'I2I3-NATS': 'I2I3-NATS',
                            'I2': 'I2',
                            'DHI-SI2': 'DHI-SI2',
                            'IDJ-I2I3': 'IDJ-I2I3',
                            'IDJ-I2NATS': 'IDJ-I2NATS',
                            'BTCH': 'BUTTON PERAL CHINA',
                            'PLBT': 'BUTTON PEARL',
                            'ANJ-I2': 'ANJ-I2',
                            'BDL-SI': 'BDL-SI',
                            'ALP-K-VS': 'ALP-K-VS',
                            'AJP-SI': 'AJP-SI',
                            'SGI-SI3': 'SGI-SI3',
                            'NMB-1-VS': 'NMB-1-VS',
                            'NMB-3-VS': 'NMB-3-VS',
                            'NMB-7-VS': 'NMB-7-VS',
                            'ALP-A-VS': 'ALP-A-VS',
                            'NMB-5-VS': 'NMB-5-VS',
                            'ALP-Y-VS': 'ALP-Y-VS',
                            'TL': 'Tourmalines',
                            'SI2-NTS': 'SI2-NTS',
                            'LUX-I3': 'LUX-I3',
                            'I1I2-NATTS': 'I1I2-NATTS',
                            'ANJ-I1I2': 'ANJ-I1I2',
                            'GSP': 'Glass Filled Sapphire',
                            'DG-I1': 'DG-I1',
                            'DG-I1-NATTS': 'DG-I1-NATTS',
                            'ANJ-SI2': 'ANJ-SI2',
                            'I3P': 'I3P',
                            'I3SCP': 'I3SCP',
                            'DHI-I1I2-NATTS': 'DHI-I1I2-NATTS',
                            'DHI-I1I2': 'DHI-I1I2',
                            'VS1': 'VS1',
                            'VDC-SI': 'VDC-SI',
                            'DG-SI': 'DG-SI',
                            'SGI-I1': 'SGI-I1',
                            'BDL-SI3-I1': 'BDL-SI3-I1',
                            'DHI-SI': 'DHI-SI',
                            'LUX-STR': 'LUX-STR',
                            'RU-AA': 'Ruby AA',
                            'VVS': 'VVS',
                            'AJP-SI2': 'AJP-SI2',
                            'LSP': 'Lab Sapphire',
                            'MG': 'MORGANITE',
                            'CH': 'CHINA PERAL',
                            'RU-OP': 'RUBY-OPAQUE',
                            'ANJ-I2I3': 'ANJ-I2I3',
                            'AJP-I2I3': 'AJP-I2I3',
                            'TP': 'Topaz',
                            'CHMPG': 'Champagne',
                            'HKIU-VS': 'HKIU-VS',
                            'AYI-NATTS': 'AYI-NATTS',
                            'AGT': 'Agate',
                            'AB': 'Abalone',
                            'OX': 'Onyx',
                            'HRG-SI2': 'HRG-SI2',
                            'AJP-VS2': 'AJP-VS2',
                            'LUX-VS': 'LUX-VS',
                            'EM-AA': 'Emerald AA',
                            'RU-AAA': 'Ruby AAA',
                            'HKIU-SI': 'HKIU-SI',
                            'HKIU2-VS': 'HKIU2-VS',
                            'HKIU2-SI2': 'HKIU2-SI2',
                            'I1-NATTS': 'I1-NATTS',
                            'FSA-SI': 'FSA-SI',
                            'EM-A': 'Emerald A',
                            'OBU-VVS': 'OBU-VVS',
                            'LUX-I3SCP': 'LUX-I3SCP',
                            'ROL-SII1': 'ROL-SII1',
                            'ROL-I1I2': 'ROL-I1I2',
                            'SP-AAA': 'Sapphire AAA',
                            'CLG-VS': 'CLG-VS',
                            'DG-VS': 'DG-VS',
                            'HKD-SI': 'HKD-SI',
                            'SI-INTYL': 'SI-INTYL',
                            'MOP': 'Mother Of Pearl',
                            'NLL-VS2': 'NLL-VS2',
                            'NLL-SI2': 'NLL-SI2',
                            'OPL-CBH': 'Opal Cabuchone',
                            'TP-AA': 'Topaz AA',
                            'GRB': 'Ruby Glass Filled',
                            'I2I3-': 'I2I3-',
                            'DCT-STR': 'DCT-STR',
                            'QTZ-CPG': 'Quartz Champagne',
                            'AM-CBH': 'Amethysts Cabuchone',
                            'GR-VS': 'GR-VS',
                            'OBU-VS': 'OBU-VS',
                            'AQM-AAA': 'Aquamarine AAA',
                            'MG-AAA': 'Morganite-AAA',
                            'TZ-AAA': 'Tanzanite AAA',
                            'OBU-STR': 'OBU-STR',
                            'OBU-SI': 'OBU-SI',
                            'BL-VS': 'BL-VS',
                            'VER-I1-TTLC': 'VER-I1-TTLC',
                            'AQM-CBH': 'AQM-CBH',
                            'PD-CBH': 'PD-CBH',
                            'TZ-AA': 'TZ-AA',
                            'CBH-MS': 'Cabuchone Moonstone',
                            'MS': 'MOON STONE',
                            'OPL': 'Opal',
                            'QTZ': 'Quartz',
                            'OBU-VS2': 'OBU-VS2',
                            'CH-DG': 'CHINA PEARL DRILL',
                            'NGL-VS': 'NGL-VS',
                            'AMD-I1': 'AMD-I1',
                            'JD1-VS': 'JD1-VS',
                            'JD1-SI': 'JD1-SI',
                            'MC': 'Malachite',
                            'RBL-VS': 'RBL-VS',
                            'SP-A': 'Sapphire A',
                            'SVK-VS': 'SVK-VS',
                            'SVK-SI': 'SVK-SI',
                            'JD1-I1I2': 'JD1-I1I2',
                            'TP-CBH': 'TP-CBH',
                            'SVK-I1': 'SVK-I1',
                            'SVK-I2': 'SVK-I2',
                            'SVK-I3': 'SVK-I3',
                            'SMQ-CBH': 'SMQ-CBH',
                            'LPLZ': 'Lapiz Lazuli',
                            'MWI-SI': 'MWI-SI',
                            'DG': 'DG',
                            'JJL-I2': 'JJL-I2',
                            'RH': 'Rhodolite Garnet',
                            'MGW-VS': 'MGW-VS',
                            'MGW-I2I3': 'MGW-I2I3',
                            'MGW-SI': 'MGW-SI',
                            'MGW': 'MGW',
                            'SMQ-AA': 'Smokey Quartz AA',
                            'CIT-AA': 'Citrine AA',
                            'LONDON-TP-AA': 'London Topoz AA',
                            'VER-VS': 'VER-VS',
                            'EM-AAA': 'EM-AAA',
                            'JD1-DG-I1': 'JD1-DG-I1',
                            'SHF-VS': 'SHF-VS',
                            'CAN-TP-AA': 'Canary Topaz',
                            'GRP-TP-AA': 'Grape Topaz',
                            'IMP-TP-AA': 'Imperial Topaz',
                            'KIWI-TP-AA': 'Kiwi Topaz',
                            'OC-TP-AA': 'Ocean Topaz-AA',
                            'SUNSET-TP-AA': 'Sunset Topaz',
                            'SKBL-TP-AA': 'Sky Blue Topaz',
                            'FSA-VS': 'FSA-VS',
                            'SMQ-CBH-AA': 'SMQ-CBH-AA',
                            'AM-CBH-AA': 'AM-CBH-AA',
                            'LNDN-TP-CBH-AA': 'LNDN-TP-CBH-AA',
                            'AJP-I1': 'AJP-I1',
                            'TRM': 'Tourmaline',
                            'VER-SI': 'VER-SI',
                            'PD-AA': 'Peridot-AA',
                            'TL-AA': 'Tourmalines-AA'
                        }
                        
                        # Extract base quality code (e.g., "EM-AA" -> "EM")
                        base_code = qlty_code_cs.split('-')[0] if '-' in qlty_code_cs else qlty_code_cs
                        stone_name = stone_name_map.get(base_code, qlty_code_cs)
                        grouped_data[style_id]['RmName_CS'] = stone_name
                    
                    # Set SetCode and SetName for Tray
                    set_code = row_dict.get('setcode')
                    set_name = row_dict.get('setname')
                    if set_code and not grouped_data[style_id]['SetCode']:
                        grouped_data[style_id]['SetCode'] = set_code
                    if set_name and not grouped_data[style_id]['SetName']:
                        grouped_data[style_id]['SetName'] = set_name
        
        # Function to format color stones for display
        def format_color_stones(color_stones_dict):
            """Format color stones as comma-separated 'QltyCode-Pieces'"""
            if not color_stones_dict:
                return "N/A"
            
            formatted_stones = []
            for qlty_code, pieces in color_stones_dict.items():
                formatted_stones.append(f"{qlty_code}-{pieces}")
            
            return ", ".join(formatted_stones)
        
        # Process the grouped data to format color stones
        for style_id, item in grouped_data.items():
            if item['StyleCode'] and item['ColorStones']:
                item['FormattedColorStones'] = format_color_stones(item['ColorStones'])
            else:
                item['FormattedColorStones'] = "N/A"
        
        # Filter out entries without a StyleCode and return
        result = [item for item in grouped_data.values() if item['StyleCode']]
        return result
    
    except Exception as e:
        print(f"Error fetching catalogue data: {e}")
        return None
    finally:
        if conn:
            conn.close()


@app.route('/api/catalogue-data', methods=['POST'])
def get_catalogue_data_api():
    """API endpoint to fetch catalogue data"""
    data = request.get_json()
    style_codes = data.get('style_codes', [])
    
    if not style_codes:
        return jsonify({'error': 'No style codes provided'}), 400
    
    catalogue_data = get_catalogue_data(style_codes)
    
    if catalogue_data is None:
        return jsonify({'error': 'Database error'}), 500
    
    return jsonify({'data': catalogue_data})

@app.route('/api/upload-excel', methods=['POST'])
def upload_excel():
    """Process uploaded Excel file and extract style codes and tray numbers"""
    try:
        if 'excel_file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['excel_file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file'}), 400
        
        # Read Excel file
        df = pd.read_excel(file)
        
        # Validate required columns
        required_columns = ['StyleCode', 'TrayNo', 'TraySRNo.']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({'error': f'Missing required columns: {", ".join(missing_columns)}'}), 400
        
        # Convert DataFrame to list of dictionaries
        data = df.to_dict('records')
        
        # Clean data - remove rows with empty StyleCode
        cleaned_data = []
        for row in data:
            if pd.notna(row.get('StyleCode')) and str(row['StyleCode']).strip():
                cleaned_data.append({
                    'StyleCode': str(row['StyleCode']).strip(),
                    'TrayNo': str(row.get('TrayNo', '')).strip() if pd.notna(row.get('TrayNo')) else '',
                    'TraySRNo': str(row.get('TraySRNo.', '')).strip() if pd.notna(row.get('TraySRNo.')) else ''
                })
        
        return jsonify({'data': cleaned_data})
        
    except Exception as e:
        return jsonify({'error': f'Error processing Excel file: {str(e)}'}), 500

@app.route('/api/catalogue-data-with-tray', methods=['POST'])
def get_catalogue_data_with_tray_api():
    """API endpoint to fetch catalogue data with tray information from Excel"""
    data = request.get_json()
    style_codes = data.get('style_codes', [])
    excel_data = data.get('excel_data', [])
    
    if not style_codes:
        return jsonify({'error': 'No style codes provided'}), 400
    
    # Get catalogue data
    catalogue_data = get_catalogue_data(style_codes)
    
    if catalogue_data is None:
        return jsonify({'error': 'Database error'}), 500
    
    # Merge with tray information from Excel
    excel_tray_info = {row['StyleCode']: row for row in excel_data}
    
    for item in catalogue_data:
        style_code = item.get('StyleCode')
        if style_code in excel_tray_info:
            tray_info = excel_tray_info[style_code]
            item['TrayNo'] = tray_info.get('TrayNo', '')
            item['TraySRNo'] = tray_info.get('TraySRNo', '')
        else:
            item['TrayNo'] = ''
            item['TraySRNo'] = ''
    
    return jsonify({'data': catalogue_data})

@app.route('/generate-pdf/<layout>')
def generate_pdf_catalogue(layout):
    """Generate PDF catalogue in specified layout (2x2, 3x3, 4x4)"""
    style_codes = request.args.get('style_codes', '').split(',')
    style_codes = [code.strip() for code in style_codes if code.strip()]
    
    # Get Excel tray data if available
    excel_data_json = request.args.get('excel_data', '')
    excel_data = []
    if excel_data_json:
        try:
            import json
            excel_data = json.loads(excel_data_json)
        except:
            excel_data = []
    
    if not style_codes:
        return jsonify({'error': 'No style codes provided'}), 400
    
    # Get catalogue data
    data = get_catalogue_data(style_codes)
    if not data:
        return jsonify({'error': 'No data found'}), 404
    
    # Merge with tray information from Excel if available
    if excel_data:
        excel_tray_info = {row['StyleCode']: row for row in excel_data}
        for item in data:
            style_code = item.get('StyleCode')
            if style_code in excel_tray_info:
                tray_info = excel_tray_info[style_code]
                item['TrayNo'] = tray_info.get('TrayNo', '')
                item['TraySRNo'] = tray_info.get('TraySRNo', '')
            else:
                item['TrayNo'] = ''
                item['TraySRNo'] = ''
    
    # Determine grid size
    grid_map = {'2x2': 2, '3x3': 3, '4x4': 4}
    grid_size = grid_map.get(layout, 2)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4) if grid_size > 2 else A4,
                           topMargin=0.5*cm, bottomMargin=0.5*cm,
                           leftMargin=0.5*cm, rightMargin=0.5*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Process data in chunks
    items_per_page = grid_size * grid_size
    for page_start in range(0, len(data), items_per_page):
        page_data = data[page_start:page_start + items_per_page]
        
        # Create grid
        table_data = []
        for row_idx in range(grid_size):
            row = []
            for col_idx in range(grid_size):
                item_idx = row_idx * grid_size + col_idx
                if item_idx < len(page_data):
                    item = page_data[item_idx]
                    cell_content = create_pdf_cell(item)
                    row.append(cell_content)
                else:
                    row.append("")
            table_data.append(row)
        
        # Create table
        col_width = (landscape(A4)[0] - 1*cm) / grid_size if grid_size > 2 else (A4[0] - 1*cm) / grid_size
        row_height = (landscape(A4)[1] - 1*cm) / grid_size if grid_size > 2 else (A4[1] - 1*cm) / grid_size
        
        table = Table(table_data, colWidths=[col_width]*grid_size, rowHeights=[row_height]*grid_size)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        
        elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f'catalogue_{layout}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

def create_pdf_cell(item):
    """Create cell content for PDF with image and text"""
    from reportlab.platypus import Paragraph, Table, TableStyle
    from reportlab.lib.styles import ParagraphStyle
    
    # Create a table with image on top and text below
    cell_data = []
    
    # Search for image dynamically based on style code
    image_path = None
    style_code = item.get('StyleCode')
    
    if style_code:
        # First try the database paths if they exist
        for path_key in ['ImagePath', 'Img3DPath', 'ImgLDPath', 'ImgPngPath']:
            if item.get(path_key) and os.path.exists(item[path_key]):
                image_path = item[path_key]
                break
        
        # If no database path found, search dynamically
        if not image_path:
            image_path = find_image_by_style_code(style_code)
    
    if image_path and os.path.exists(image_path):
        try:
            # Create image with proper sizing
            img = RLImage(image_path)
            img.drawWidth = 2.5 * cm  # Width of image
            img.drawHeight = 2.5 * cm  # Height of image
            cell_data.append([img])
        except Exception as e:
            print(f"Error adding image to PDF: {e}")
            # Add placeholder if image fails
            cell_data.append(["[Image Not Available]"])
    else:
        # Add placeholder if no image found
        cell_data.append(["[Image Not Available]"])
    
    # Add text content below image
    style = ParagraphStyle(
        'CustomStyle',
        fontSize=8,
        leading=10,
    )
    
    dia_details = f"{item['Pieces_D']} Qty / {item['Weight_D']:.2f} Cts" if item['Pieces_D'] else "N/A"
    col_qty = item.get('FormattedColorStones', 'N/A')
    grams = f"{item['Weight_M']:.3f}" if item['Weight_M'] else "N/A"
    
    # Get tray information
    tray_info = ""
    if item.get('TrayNo') and item.get('TraySRNo'):
        tray_info = f"{item['TrayNo']} - {item['TraySRNo']}"
    elif item.get('TrayNo'):
        tray_info = item['TrayNo']
    else:
        tray_info = "N/A"
    
    text = f"""<b>Design No:</b> {item['StyleCode']}<br/>
<b>Grams:</b> {grams}<br/>
<b>Dia Details:</b> {dia_details}<br/>
<b>Col Qty:</b> {col_qty}<br/>
<b>Tray:</b> {tray_info}"""
    
    cell_data.append([Paragraph(text, style)])
    
    # Create table for this cell
    cell_table = Table(cell_data, colWidths=[3*cm])
    cell_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    
    return cell_table

@app.route('/generate-excel/<layout>')
def generate_excel_catalogue(layout):
    """Generate Excel catalogue in specified layout (2x2, 3x3, 4x4) matching the exact format from image"""
    style_codes = request.args.get('style_codes', '').split(',')
    style_codes = [code.strip() for code in style_codes if code.strip()]
    
    # Get Excel tray data if available
    excel_data_json = request.args.get('excel_data', '')
    excel_data = []
    if excel_data_json:
        try:
            import json
            excel_data = json.loads(excel_data_json)
        except:
            excel_data = []
    
    if not style_codes:
        return jsonify({'error': 'No style codes provided'}), 400
    
    # Get catalogue data
    data = get_catalogue_data(style_codes)
    if not data:
        return jsonify({'error': 'No data found'}), 404
    
    # Merge with tray information from Excel if available
    if excel_data:
        excel_tray_info = {row['StyleCode']: row for row in excel_data}
        for item in data:
            style_code = item.get('StyleCode')
            if style_code in excel_tray_info:
                tray_info = excel_tray_info[style_code]
                item['TrayNo'] = tray_info.get('TrayNo', '')
                item['TraySRNo'] = tray_info.get('TraySRNo', '')
            else:
                item['TrayNo'] = ''
                item['TraySRNo'] = ''
    
    # Determine grid size
    grid_map = {'2x2': 2, '3x3': 3, '4x4': 4}
    grid_size = grid_map.get(layout, 2)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Catalogue {layout}"
    
    # Each item will be a 2x6 table (2 columns: labels and values, 6 rows: Image + 5 fields)
    # Plus 1 blank column between items and 1 blank row after each item
    # Calculate dimensions
    item_width_cols = 2  # 2 columns per item (label, value)
    blank_col_width = 1  # 1 blank column between items
    item_height_rows = 6  # 6 rows per item (image row + 5 data rows)
    blank_row_height = 1  # 1 blank row after each item row
    
    # Total columns = (items × 2) + (gaps × 1) - 1
    # For 2x2: (2 × 2) + (2 - 1) = 5 columns per row
    # For 3x3: (3 × 2) + (3 - 1) = 8 columns per row
    total_cols = (grid_size * item_width_cols) + (grid_size - 1)
    
    # Set column widths
    for col_idx in range(total_cols):
        col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
        # Calculate position in pattern: Label, Value, Blank, Label, Value, Blank...
        position_in_pattern = col_idx % 3
        if position_in_pattern == 0:  # Label columns (0, 3, 6, 9...)
            ws.column_dimensions[col_letter].width = 15
        elif position_in_pattern == 1:  # Value columns (1, 4, 7, 10...)
            ws.column_dimensions[col_letter].width = 20
        else:  # Blank columns (2, 5, 8, 11...)
            ws.column_dimensions[col_letter].width = 2
    
    # Process data
    items_per_page = grid_size * grid_size
    current_excel_row = 1
    
    for page_start in range(0, len(data), items_per_page):
        page_data = data[page_start:page_start + items_per_page]
        
        # Process each grid row (e.g., for 2x2: 2 rows of items)
        for grid_row_idx in range(grid_size):
            start_row = current_excel_row
            
            # Set row heights for this item row
            for row_offset in range(item_height_rows):
                row_num = start_row + row_offset
                if row_offset == 0:  # Image row
                    ws.row_dimensions[row_num].height = 120
                else:  # Data rows
                    ws.row_dimensions[row_num].height = 25
            
            # Set blank row height after this item row
            blank_row_num = start_row + item_height_rows
            ws.row_dimensions[blank_row_num].height = 15  # Blank row for spacing
            
            # Process each item in this grid row (e.g., for 2x2: 2 items per row)
            for grid_col_idx in range(grid_size):
                item_idx = grid_row_idx * grid_size + grid_col_idx
                if item_idx >= len(page_data):
                    continue
                
                item = page_data[item_idx]
                
                # Calculate column positions for this item
                # Each item takes 2 columns, plus 1 blank column after (except last item)
                # Item 0: cols 1-2, blank 3
                # Item 1: cols 4-5, blank 6
                # Item 2: cols 7-8, blank 9
                col_start = grid_col_idx * (item_width_cols + blank_col_width) + 1
                col_label = col_start
                col_value = col_start + 1
                
                # Row 1: Image (merged across both columns)
                img_cell = ws.cell(row=start_row, column=col_start)
                img_cell.value = "Image"
                img_cell.alignment = Alignment(horizontal='center', vertical='center')
                img_cell.fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
                img_cell.font = Font(bold=True, size=10, color='4472C4')
                ws.merge_cells(start_row=start_row, start_column=col_start, 
                              end_row=start_row, end_column=col_value)
                
                # Try to add image if path exists
                image_path = None
                style_code = item.get('StyleCode')
                
                if style_code:
                    # First try the database paths if they exist
                    for path_key in ['ImagePath', 'Img3DPath', 'ImgLDPath', 'ImgPngPath']:
                        if item.get(path_key) and os.path.exists(item[path_key]):
                            image_path = item[path_key]
                            break
                    
                    # If no database path found, search dynamically
                    if not image_path:
                        image_path = find_image_by_style_code(style_code)
                
                if image_path and os.path.exists(image_path):
                    try:
                        img = XLImage(image_path)
                        # Resize image to fit
                        img.width = 200
                        img.height = 150
                        ws.add_image(img, f"{openpyxl.utils.get_column_letter(col_start)}{start_row}")
                    except Exception as e:
                        print(f"Error adding image to Excel: {e}")
                
                # Row 2: Design No
                row_num = start_row + 1
                label_cell = ws.cell(row=row_num, column=col_label)
                label_cell.value = "Design No :"
                label_cell.font = Font(bold=True, size=10)
                label_cell.alignment = Alignment(horizontal='left', vertical='center')
                label_cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                value_cell = ws.cell(row=row_num, column=col_value)
                value_cell.value = item.get('StyleCode', 'N/A')
                value_cell.font = Font(size=10, color='0000FF')
                value_cell.alignment = Alignment(horizontal='center', vertical='center')
                value_cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # Row 3: Grams
                row_num = start_row + 2
                label_cell = ws.cell(row=row_num, column=col_label)
                label_cell.value = "Grams :"
                label_cell.font = Font(bold=True, size=10)
                label_cell.alignment = Alignment(horizontal='left', vertical='center')
                label_cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                value_cell = ws.cell(row=row_num, column=col_value)
                weight_m = item.get('Weight_M', 0)
                value_cell.value = f"{weight_m:.3f}" if weight_m else "N/A"
                value_cell.alignment = Alignment(horizontal='center', vertical='center')
                value_cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # Row 4: Dia Details
                row_num = start_row + 3
                label_cell = ws.cell(row=row_num, column=col_label)
                label_cell.value = "Dia Details :"
                label_cell.font = Font(bold=True, size=10)
                label_cell.alignment = Alignment(horizontal='left', vertical='center')
                label_cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                value_cell = ws.cell(row=row_num, column=col_value)
                pieces_d = item.get('Pieces_D', 0)
                weight_d = item.get('Weight_D', 0)
                value_cell.value = f"{pieces_d} Qty / {weight_d:.2f} Cts" if pieces_d else "N/A"
                value_cell.alignment = Alignment(horizontal='center', vertical='center')
                value_cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # Row 5: Col Qty
                row_num = start_row + 4
                label_cell = ws.cell(row=row_num, column=col_label)
                label_cell.value = "Col Qty :"
                label_cell.font = Font(bold=True, size=10)
                label_cell.alignment = Alignment(horizontal='left', vertical='center')
                label_cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                value_cell = ws.cell(row=row_num, column=col_value)
                value_cell.value = item.get('FormattedColorStones', 'N/A')
                value_cell.alignment = Alignment(horizontal='center', vertical='center')
                value_cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # Row 6: Tray
                row_num = start_row + 5
                label_cell = ws.cell(row=row_num, column=col_label)
                label_cell.value = "Tray :"
                label_cell.font = Font(bold=True, size=10)
                label_cell.alignment = Alignment(horizontal='left', vertical='center')
                label_cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                value_cell = ws.cell(row=row_num, column=col_value)
                # Use tray information from Excel data if available, otherwise use database SetCode/SetName
                tray_no = item.get('TrayNo', '')
                tray_sr_no = item.get('TraySRNo', '')
                
                if tray_no and tray_sr_no:
                    value_cell.value = f"{tray_no} - {tray_sr_no}"
                elif tray_no:
                    value_cell.value = tray_no
                else:
                    # Fallback to database tray information
                    set_code = item.get('SetCode', '')
                    set_name = item.get('SetName', '')
                    if set_code and set_name:
                        value_cell.value = f"{set_code} - {set_name}"
                    elif set_code:
                        value_cell.value = set_code
                    else:
                        value_cell.value = "N/A"
                value_cell.alignment = Alignment(horizontal='center', vertical='center')
                value_cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            
            # Move to next row of items (including blank row)
            current_excel_row += item_height_rows + blank_row_height
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f'catalogue_{layout}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(buffer, as_attachment=True, download_name=filename, 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True,host='0.0.0.0')